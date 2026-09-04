"""Read-only scope check: compare the exported CPI note edit with pre-turn HEAD."""
import io, json, subprocess, sys
from pathlib import Path
from zipfile import ZipFile
from openpyxl import load_workbook
HERE=Path(__file__).resolve().parent
ROOT=Path(subprocess.check_output(['git','rev-parse','--show-toplevel'],cwd=HERE,text=True,encoding='utf-8').strip())
canonical=HERE.parent/'SMARTBI_TYPE_AUDIT_V50.xlsx'
out=Path(sys.argv[1]) if len(sys.argv)>1 else canonical
old_bytes=subprocess.check_output(['git','show',f'f73b38828082f74b13a0c5bdb878d4186894203c:{canonical.relative_to(ROOT).as_posix()}'],cwd=ROOT)
new_bytes=out.read_bytes()
old,new=[load_workbook(io.BytesIO(b),data_only=False) for b in [old_bytes,new_bytes]]
assert old.sheetnames==new.sheetnames
changes=[]
for name in old.sheetnames:
    a,b=old[name],new[name]
    assert (a.max_row,a.max_column)==(b.max_row,b.max_column)
    for prop in ['merged_cells','freeze_panes','data_validations','sheet_state']:
        assert str(getattr(a,prop))==str(getattr(b,prop)),prop
    assert a.auto_filter.ref==b.auto_filter.ref
    assert list(a.tables)==list(b.tables)
    for row in a:
        for c in row:
            d=b[c.coordinate]
            if c.value!=d.value:
                assert name=='类型审计' and c.coordinate=='F88'
                changes.append(c.coordinate)
            for prop in ['font','fill','border','alignment','number_format','protection']:
                assert str(getattr(c,prop))==str(getattr(d,prop)),(c.coordinate,prop)
            assert c.comment==d.comment
    for key in set(a.column_dimensions)|set(b.column_dimensions):
        assert dict(a.column_dimensions[key])==dict(b.column_dimensions[key])
    for key in set(a.row_dimensions)|set(b.row_dimensions):
        left,right=dict(a.row_dimensions[key]),dict(b.row_dimensions[key])
        if key==88:
            left.pop('ht',None);right.pop('ht',None)
        assert left==right,(key,left,right)
assert changes==['F88']
counts={}
for row in new.worksheets[0].iter_rows(min_row=2,values_only=True):
    if row[0] and row[1]:counts[row[4]]=counts.get(row[4],0)+1
assert counts=={'PASS':383,'FAIL':59}
with ZipFile(io.BytesIO(old_bytes)) as a,ZipFile(io.BytesIO(new_bytes)) as b:
    assert set(a.namelist())==set(b.namelist())
    changed_parts=[p for p in a.namelist() if a.read(p)!=b.read(p)]
    assert set(changed_parts)<={'xl/worksheets/sheet1.xml','xl/workbook.xml','_rels/.rels','xl/_rels/workbook.xml.rels'}
print(json.dumps({'result':'PASS','changedCells':changes,'counts':counts,'stylesAndFeaturesPreserved':True,'changedParts':changed_parts},ensure_ascii=False,indent=2))
