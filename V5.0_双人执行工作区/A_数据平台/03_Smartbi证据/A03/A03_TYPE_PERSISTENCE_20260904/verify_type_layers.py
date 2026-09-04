"""Read-only source/XML extraction; write sanitized evidence, never modify XLSX/XML.

Run with the bundled Python runtime. Historical XML is local-only and ignored by Git.
The 59 rows remain calculation/review pending even if persisted types agree.
"""
from collections import Counter
from decimal import Decimal
from hashlib import sha256
from pathlib import Path
import csv
import json
import subprocess
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
A03 = HERE.parent
EVIDENCE = A03.parent
MODEL_ID = "6b5dff57a4093ba3db07d2903905fe40"
XML = EVIDENCE / "P1/A_DB04_MODEL_METADATA_PATCH_20260901/migrate_MODEL_METADATA_PATCH_20260901.xml"
PREFLIGHT = A03 / "A03_RECHECK_20260904/A03_WORKBOOK_PREFLIGHT_20260904.json"
SOURCE = EVIDENCE.parent / "01_输入只读镜像/D0-D12_数据交付_V4.2/data/smartbi/country_monthly_risk.xlsx"
LIVE = HERE / "LIVE_TYPE_TABLES.json"
TYPE_BOOK = A03 / "SMARTBI_TYPE_AUDIT_V50.xlsx"


def digest(path):
    return sha256(path.read_bytes()).hexdigest()


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8-sig"))


def one(items, description):
    assert len(items) == 1, f"Ambiguous/missing {description}: {len(items)}"
    return items[0]


def compatible(expected, actual):
    return actual in {
        "数值": {"DOUBLE", "BIGDECIMAL", "INTEGER", "LONG", "FLOAT"},
        "整数": {"INTEGER", "LONG"},
        "日期": {"DATE", "DATETIME", "TIMESTAMP"},
        "三态": {"INTEGER", "LONG"},
    }[expected]


def main():
    book_hash_before = digest(TYPE_BOOK)
    repo = Path(subprocess.check_output(["git", "rev-parse", "--show-toplevel"], cwd=HERE, text=True, encoding="utf-8").strip())
    baseline_commit = "6c49f46ec907892d9e44820bc40511803f9cd9ca"
    baseline_bytes = subprocess.check_output(["git", "show", f"{baseline_commit}:{TYPE_BOOK.relative_to(repo).as_posix()}"], cwd=repo)
    ds = one([d for d in ET.parse(XML).getroot().findall("AUGMENTED_DATASET")
              if d.get("id") == MODEL_ID or d.findtext("id") == MODEL_ID], "formal model")
    views = json.loads(ds.findtext("define"))["views"]
    # Whitelist only field metadata. Never emit dataSource, define, credentials or creator IDs.
    model_fields = [dict(f.attrib) for f in ds.find("fields")]
    exceptions = read_json(PREFLIGHT)["typeRecordedExceptions"]
    live = read_json(LIVE)
    assert live["modelId"] == MODEL_ID
    assert len(exceptions) == 59
    expected_ui = {"STRING": "字符串", "DOUBLE": "浮点型", "BIGDECIMAL": "长浮点型",
                   "INTEGER": "整型", "DATE": "日期"}
    results = []
    for item in exceptions:
        object_name = "V50_" + Path(item["table"]).stem
        view = one([v for v in views if v.get("alias", "").casefold() == object_name.casefold()], object_name)
        source_field = one([f for f in view["fields"] if f["name"] == item["field"]], item["field"])
        model_field = one([f for f in model_fields if f.get("viewId") == view["id"]
                           and f.get("referenceFieldId") == source_field["id"]], "model field")
        live_table = one([t for t in live["tables"] if t["table"].casefold() == object_name.casefold()], "live table")
        live_row = one([r for r in live_table["rows"][1:] if r[0] == model_field["name"]], "live field")
        ui_agrees = live_row[2] == expected_ui[model_field["valueType"]]
        type_agrees = compatible(item["expected"], model_field["valueType"])
        results.append({
            "table": item["table"], "field": item["field"], "expected": item["expected"],
            "historicalLedgerStatus": item["recordedStatus"],
            "sourceViewType_20260901": source_field["valueType"],
            "persistedModelType_20260901": model_field["valueType"],
            "persistedModelFieldName": model_field["name"],
            "liveModelType_20260904": live_row[2],
            "modelTypeCompatible": type_agrees, "liveMatchesPersisted": ui_agrees,
            "calculationAcceptance": "PENDING", "bReview": "PENDING",
        })
    assert all(r["modelTypeCompatible"] for r in results)
    assert all(r["liveMatchesPersisted"] for r in results), [r for r in results if not r["liveMatchesPersisted"]]
    assert digest(TYPE_BOOK) == book_hash_before
    report = {
        "checkedAt": live["capturedAt"],
        "modelId": MODEL_ID, "modelAlias": "MDL_XH202612_V50_COUNTRY_RESERVE",
        "scope": "59 historical exceptions only. Existing 2026-09-01 export plus 2026-09-04 read-only UI. Not a fresh export, calculation PASS or B signoff.",
        "xmlFile": XML.name, "xmlSha256": digest(XML),
        "xmlModelLastModified": ds.get("lastModified") or ds.findtext("lastModified"),
        "liveEvidenceFile": LIVE.name, "liveEvidenceSha256": digest(LIVE),
        "historicalLedgerCommit": baseline_commit,
        "historicalLedgerSha256Unchanged": sha256(baseline_bytes).hexdigest(),
        "counts": {
            "historicalExceptions": len(results),
            "persistedTypeCompatible": sum(r["modelTypeCompatible"] for r in results),
            "liveTypeMatchesPersisted": sum(r["liveMatchesPersisted"] for r in results),
            "sourceViewTypes": dict(Counter(r["sourceViewType_20260901"] for r in results)),
            "persistedModelTypes": dict(Counter(r["persistedModelType_20260901"] for r in results)),
            "calculationAcceptancePending": len(results), "bReviewPending": len(results),
        }, "fields": results,
    }
    (HERE / "TYPE_LAYER_RECONCILIATION.json").write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8", newline="\n")
    with (HERE / "TYPE_LAYER_RECONCILIATION.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(results[0]), lineterminator="\n")
        writer.writeheader()
        writer.writerows(results)

    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    cpi = headers.index("cpi_index")
    values, row_count, invalid = [], 0, []
    for n, row in enumerate(rows, 2):
        row_count += 1
        value = row[cpi]
        if value is None:
            continue
        if not isinstance(value, (int, float)) or isinstance(value, bool):
            invalid.append(n)
        else:
            values.append(Decimal(str(value)))
    workbook.close()
    assert row_count == 7680 and len(values) == 7616 and not invalid
    baseline = {
        "scope": "Source-only full-table CPI baseline. Never substitute these values for platform results.",
        "sourceFile": SOURCE.name, "sourceSha256": digest(SOURCE), "field": "cpi_index",
        "rows": row_count, "nonNull": len(values), "null": row_count - len(values),
        "min": str(min(values)), "max": str(max(values)),
        "sum": str(sum(values)), "average": str(sum(values) / len(values)),
        "numericMethod": "Decimal(str(openpyxl numeric cell value)); ignore None only, no fill-zero, no distinct",
        "platformAggregate": None, "calculationStatus": "PENDING",
    }
    (HERE / "CPI_SOURCE_BASELINE.json").write_text(json.dumps(baseline, ensure_ascii=False, indent=2) + "\n", encoding="utf-8", newline="\n")
    manifest = [{"file": p.name, "bytes": p.stat().st_size, "sha256": digest(p)}
                for p in sorted(HERE.iterdir()) if p.is_file() and p.name != "EVIDENCE_MANIFEST.json"]
    (HERE / "EVIDENCE_MANIFEST.json").write_text(json.dumps({"files": manifest}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8", newline="\n")
    print(json.dumps({"counts": report["counts"], "sourceBaseline": baseline}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
