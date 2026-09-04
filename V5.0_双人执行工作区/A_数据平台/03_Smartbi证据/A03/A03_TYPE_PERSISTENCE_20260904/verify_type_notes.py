"""Read-only workbook scope verification against the immutable pre-turn commit."""
import io
import json
from pathlib import Path
import subprocess
import sys
import xml.etree.ElementTree as ET
from zipfile import ZipFile

from openpyxl import load_workbook

here = Path(__file__).resolve().parent
repo = Path(subprocess.check_output(["git", "rev-parse", "--show-toplevel"], cwd=here, text=True, encoding="utf-8").strip())
canonical = here.parent / "SMARTBI_TYPE_AUDIT_V50.xlsx"
output = Path(sys.argv[1]) if len(sys.argv) > 1 else canonical
baseline = "6c49f46ec907892d9e44820bc40511803f9cd9ca"
old_bytes = subprocess.check_output(["git", "show", f"{baseline}:{canonical.relative_to(repo).as_posix()}"], cwd=repo)
new_bytes = output.read_bytes()
old, new = [load_workbook(io.BytesIO(b), data_only=False) for b in [old_bytes, new_bytes]]
log = json.loads((here / "TYPE_NOTES_CHANGELOG.json").read_text(encoding="utf-8"))
allowed = {c["cell"] for c in log["changes"]}
assert len(allowed) == 59
assert old.sheetnames == new.sheetnames
changes, row_height_changes = [], []
for sheet_name in old.sheetnames:
    a, b = old[sheet_name], new[sheet_name]
    assert (a.max_row, a.max_column) == (b.max_row, b.max_column)
    assert str(a.merged_cells) == str(b.merged_cells)
    assert a.freeze_panes == b.freeze_panes
    assert str(a.data_validations) == str(b.data_validations)
    assert list(a.tables) == list(b.tables)
    assert a.auto_filter.ref == b.auto_filter.ref
    for row in a:
        for cell in row:
            other = b[cell.coordinate]
            if cell.value != other.value:
                assert cell.coordinate in allowed
                changes.append(cell.coordinate)
            for prop in ["font", "fill", "border", "alignment", "number_format", "protection"]:
                assert str(getattr(cell, prop)) == str(getattr(other, prop)), (cell.coordinate, prop)
            assert cell.comment == other.comment
    for key in set(a.row_dimensions) | set(b.row_dimensions):
        if a.row_dimensions[key].height != b.row_dimensions[key].height:
            assert f"F{key}" in allowed
            row_height_changes.append(key)
    for key in set(a.column_dimensions) | set(b.column_dimensions):
        assert dict(a.column_dimensions[key]) == dict(b.column_dimensions[key]), ("column", key)
assert set(changes) == allowed
with ZipFile(io.BytesIO(old_bytes)) as a, ZipFile(io.BytesIO(new_bytes)) as b:
    assert sorted(a.namelist()) == sorted(b.namelist())
    changed_parts = [name for name in a.namelist() if a.read(name) != b.read(name)]
    assert set(changed_parts) <= {"xl/worksheets/sheet1.xml", "xl/workbook.xml", "_rels/.rels", "xl/_rels/workbook.xml.rels"}, changed_parts
    for part in ["_rels/.rels", "xl/_rels/workbook.xml.rels"]:
        left, right = [ET.fromstring(z.read(part)) for z in [a, b]]
        assert [{k:v for k,v in e.attrib.items() if k != "Id"} for e in left] == [{k:v for k,v in e.attrib.items() if k != "Id"} for e in right]
    left, right = [ET.fromstring(z.read("xl/workbook.xml")) for z in [a, b]]
    rel_key = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    for node in list(left.iter()) + list(right.iter()):
        node.attrib.pop(rel_key, None)
    assert ET.tostring(left) == ET.tostring(right)
    for archive in [a, b]:
        rels = {e.get("Id"):e.get("Target") for e in ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))}
        sheet_node = ET.fromstring(archive.read("xl/workbook.xml")).find("{*}sheets/{*}sheet")
        assert rels[sheet_node.get(rel_key)] == "/xl/worksheets/sheet1.xml"
print(json.dumps({"result": "PASS", "baseline": baseline, "changedNoteCells": len(changes),
                  "unchangedAcceptanceColumns": True, "unchangedCellStyles": True,
                  "rowHeightChangesWithinScope": len(row_height_changes), "changedXmlParts": changed_parts}, ensure_ascii=False, indent=2))
